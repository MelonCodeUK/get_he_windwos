
from flask import Flask
from flask import request
from openpyxl import load_workbook
from openpyxl import Workbook
import os


commands = []
commands_out = []
app = Flask(__name__)


@app.route('/login/<admin_name>/<password>', methods=["GET","POST"])
def login(admin_name,password):
    admin_file = load_workbook('admins_and_passwords.xlsx')
    sheet_admins = admin_file.active

    column_NAME_admins = [cell.value for cell in sheet_admins['B']]
    colum_PASSWORD_admins = [cell.value for cell in sheet_admins['C']]

    ret = None
    ret1 = None

    for value in column_NAME_admins:
        if value == admin_name:
            ret = True
            break
        else:
            ret = False
    for value in colum_PASSWORD_admins:
        if value == password:
            ret1 = True
            break
        else:
            ret1 = False
    if ret == True and ret1 == True:
        return "True"
    else:
        # print(ret,ret1)
        return "Error: wrong password or username"



@app.route('/add_admins', methods=["GET","POST"])
def add_admins():
    try:
        Inpt = str(request.args.get("info"))
        Input = Inpt.split("-")

        if os.path.exists("admins_and_passwords.xlsx"):
            admins = load_workbook('admins_and_passwords.xlsx')
            sheet_admins_number = admins.active
            colum_admins_number = [cell.value for cell in sheet_admins_number['A']]
            for j in colum_admins_number:
                pass
            print(f'A{str(int(j)+1)}')
            sheet_admins_number[f'A{int(j)+1}'] = int(j)+1
            sheet_admins_number[f'B{str(int(j)+1)}'] = Input[0]
            sheet_admins_number[f'C{str(int(j)+1)}'] = Input[1]
            admins.save("admins_and_passwords.xlsx")
        else:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '№'
            ws['B1'] = 'admin_name'
            ws['C1'] = 'admin_password'
            ws['A2'] = 1
            ws['B2'] = 'admin'
            ws['C2'] = 'admin'
            wb.save('admins_and_passwords.xlsx')
        try:
            return "True"
        except:
            return "False"
    except:
        return Exception



def checks(name_admin,password_admin,user_name):
    # print(name_admin,password_admin,user_name)
    admin_file = load_workbook('admins_and_passwords.xlsx')
    users = load_workbook('users.xlsx')
    sheet_admins = admin_file.active
    sheet_users = users.active

    column_NAME_admins = [cell.value for cell in sheet_admins['B']]
    colum_PASSWORD_admins = [cell.value for cell in sheet_admins['C']]
    colum_USERS = [cell.value for cell in sheet_users['B']]

    ret = None
    ret1 = None

    for value in column_NAME_admins:
        if value == name_admin:
            ret = True
            break
        else:
            ret = False
    for value in colum_PASSWORD_admins:
        if value == password_admin:
            ret1 = True
            break
        else:
            ret1 = False
    if ret == True and ret1 == True:
        for value in colum_USERS:
            if value == user_name:
                return True
    else:
        # print(ret,ret1)
        return "Error: wrong password or username"



def add_name(name, ip):
    if os.path.exists("users.xlsx"):
        users = load_workbook('users.xlsx')
        sheet_users_number = users.active
        colum_USERS_number = [cell.value for cell in sheet_users_number['A']]
        for j in colum_USERS_number:
            pass
        print(f'A{str(int(j)+1)}')
        sheet_users_number[f'A{int(j)+1}'] = int(j)+1
        sheet_users_number[f'B{str(int(j)+1)}'] = name
        sheet_users_number[f'C{str(int(j)+1)}'] = ip
        users.save("users.xlsx")
    else:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '№'
        ws['B1'] = 'user_name'
        ws['C1'] = 'user_ip'
        ws['A2'] = 1
        ws['B2'] = 'user_name_test'
        ws['C2'] = 'user_ip_test'
        wb.save('users.xlsx')



def chek_c(user,admin_name):
    if len(commands_out) != 0:
        for il in commands_out:
            if il[0] == user:
                if il[1] == admin_name:
                    commands_out.remove([il[0],il[1],il[2]])
                    print(il)
                    return il[2]
            else:
                return False
    else:
        return False






# http://127.0.0.1:5000/users_get?input=name_admin-password_admin-username-command_windows

@app.route('/users_get', methods=["GET","POST"])
def users_i():
    Inpt = str(request.args.get("input"))
    Input = Inpt.split("-")
    admin_name = Input[0]
    admin_pass = Input[1]
    user_name = Input[2]
    command = Input[3]



    ck = checks(admin_name, admin_pass, user_name)
    if ck == True:
        if command == "wait":
            chekcomd = chek_c(user_name,admin_name)
            if chekcomd == False:
                rents = "no_wait"
                print("ln:" + str(len(commands_out)))
            else:
                with open(chekcomd) as file_read:
                    rents = file_read.read()
                os.remove(chekcomd)

        else:
            commands.append([user_name,command,admin_name])
            rents = "no_wait"


    else:
        rents = ck
    return rents



    
def get_user(user_name):
    if os.path.exists("users.xlsx"):
        users = load_workbook('users.xlsx')
        sheet_users = users.active
        colum_USERS = [cell.value for cell in sheet_users['B']]
        for value in colum_USERS:
            if value == user_name:
               return True
            else:
                return False







@app.route('/cmd', methods=["GET","POST"])
def users_admin():
    returNs = None
    user_name = str(request.args.get("input"))
    get_u = get_user(user_name=user_name)
    if get_u == True:
        if len(commands) == 0:
            returNs = str(0)
        else:
            for i in commands:
                if 'file' in request.files and i[0] == user_name:
                    commands.remove([i[0],i[1],i[2]])
                    if os.path.exists("commands_out_AND_files"):
                        pass
                    else:
                        os.mkdir("./commands_out_AND_files")
                    print(type(request.files))
                    file = request.files['file']
                    print(file.save)
                    file.save('./commands_out_AND_files/' + i[0]+ "_" + i[1] + "_" + i[2] + "_out.txt")
                    path = './commands_out_AND_files/' + i[0]+ "_" + i[1] + "_" + i[2] + "_out.txt"
                    # print(returNs)
                    commands_out.append([i[0],i[2],path])
                    returNs = "0"
                    break
                    





                elif i[0] == user_name:
                    print("get_command")
                    returNs =  str(i[0] + "-" + i[1] + "-" + i[2])
                    break
                elif i[0] != user_name:
                    returNs = str(0)




    else:
        add_name(user_name, request.remote_addr)
        returNs = "new_name"
    print("return^" + returNs)
    return returNs


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file found', 400
    
    file = request.files['file']
    file.save('uploads/' + file.filename)
    
    return 'File saved successfully'


if __name__ == '__main__':
    app.run()
